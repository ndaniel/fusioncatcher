#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""

It downloads the fusion genes from article:
D. Robison et al., Integrative Clinical Genomics of Advanced Prostate Cancer, Cell, Vol. 161, May 2015, http://dx.doi.org/10.1016/j.cell.2015.05.001



Author: Daniel Nicorici, Daniel.Nicorici@gmail.com

Copyright (c) 2009-2021 Daniel Nicorici

This file is part of FusionCatcher.

FusionCatcher is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

FusionCatcher is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with FusionCatcher (see file 'COPYING.txt').  If not, see
<http://www.gnu.org/licenses/>.

By default, FusionCatcher is running BLAT aligner
<http://users.soe.ucsc.edu/~kent/src/> but it offers also the option to disable
all its scripts which make use of BLAT aligner if you choose explicitly to do so.
BLAT's license does not allow to be used for commercial activities. If BLAT
license does not allow to be used in your case then you may still use
FusionCatcher by forcing not use the BLAT aligner by specifying the option
'--skip-blat'. Fore more information regarding BLAT please see its license.

Please, note that FusionCatcher does not require BLAT in order to find
candidate fusion genes!

This file is not running/executing/using BLAT.
"""



import os
import sys
import urllib2
import gzip
import socket
import optparse
import concatenate
import symbols
import shutil
import datetime

if __name__ == '__main__':

    #command line parsing

    usage = "%prog [options]"
    description = """It downloads the human fusion genes found in prostate cancer patients from article: D. Robison et al., Integrative Clinical Genomics of Advanced Prostate Cancer, Cell, Vol. 161, May 2015, http://dx.doi.org/10.1016/j.cell.2015.05.001"""
    version = "%prog 0.12 beta"

    parser = optparse.OptionParser(usage=usage,description=description,version=version)

    parser.add_option("--organism","-g",
                      action = "store",
                      type = "string",
                      dest = "organism",
                      default = "homo_sapiens",
                      help="""The name of the organism for which the known fusion genes are downloaded, e.g. homo_sapiens, mus_musculus, etc. Default is '%default'.""")

    parser.add_option("--output","-o",
                      action="store",
                      type="string",
                      dest="output_directory",
                      default = '.',
                      help="""The output directory where the known fusion genes are stored. Default is '%default'.""")

    parser.add_option("--data",
                      action="store",
                      type="string",
                      dest="data_filename",
                      help="""The input Excel file containg the data from the article. It should be used when there is no internet connection to the site which hosts the article.""")


    parser.add_option("--skip-filter-overlap",
                      action="store_true",
                      dest="skip_filter_overlap",
                      default = False,
                      help="""If set then it filters out the known fusion genes where the (i) genes are fully overlapping, or (ii) the genes are partially overlapping and are on the same strand. Default is '%default'.""")


    (options,args) = parser.parse_args()

    # validate options
    if not (options.output_directory
            ):
        parser.print_help()
        sys.exit(1)

    file(os.path.join(options.output_directory,'prostate_cancer.txt'),'w').write('')

    try:
        import openpyxl
        print "The Python Excel parser OPENPYXL was found!"
    except:
        print >> sys.stderr,"WARNING: Python OPENPYXL library was not found!"
        print >> sys.stderr,"Please, install the Python OpenPYXL library in order to be able to parse the known fusion genes in prostate tumor patients!"
        sys.exit(0)


    # timeout in seconds
    timeout = 1800
    socket.setdefaulttimeout(timeout)

    # NEW:
    # http://www.sciencedirect.com/science/MiamiMultiMediaURL/1-s2.0-S0092867415005486/1-s2.0-S0092867415005486-mmc4.xlsx/272196/html/S0092867415005486/198196b9ed4de2f80a504687fdfb5126/mmc4.xlsx


    #url = 'http://www.sciencedirect.com/science/MiamiMultiMediaURL/1-s2.0-S0092867415005486/1-s2.0-S0092867415005486-mmc4.xlsx/272196/html/S0092867415005486/198196b9ed4de2f80a504687fdfb5126/mmc4.xlsx'
    url = 'https://www.ncbi.nlm.nih.gov/pmc/articles/PMC4484602/bin/NIHMS702018-supplement-Table_S4.xlsx'
    tmp_file = os.path.join(options.output_directory,'temp_prostates.xlsx')

    headers = { 'User-Agent' : 'Mozilla/5.0' }


    if options.organism.lower() == 'homo_sapiens':
        today = datetime.date.today()
        data = []
        sem = True
        if options.data_filename:
            tmp_file = options.data_filename
            print "Using the local file..."
        else:
            print "Downloading the known fusion genes from the article..."
            try:
                req = urllib2.Request('%s' % (url,), headers=headers)
                da = urllib2.urlopen(req)
                file(tmp_file,'w').write(da.read())
            except:
                print >>sys.stderr, "Warning: Cannot access '%s'! The output file will be empty!" % (url,)
                sem = False

        data = []
        if sem:
            print "Parsing..."
            # parse the ChimerDB file with the known fusion genes
            wb = None
            data = set()

            try:
                # new version of openpyxl
                print "Using the new version of openpyxl..."
                wb = openpyxl.load_workbook(filename=tmp_file)
            except:
                print "The new version of openpyxl did not work and now using the old version of openpyxl..."
                # old version of openpyxl
                try:
                    wb = openpyxl.reader.excel.load_workbook(filename=tmp_file)
                except:
                    print "WARNING: Not able to use the python openpyxl library!"
                    if not options.data_filename:
                        os.remove(tmp_file)
                    sys.exit(0)

            print "..."
            # parse the Sheet1 sheet
            sheet = wb.get_sheet_by_name(name = 'GENE FUSIONS') #

            for i,row in enumerate(sheet.rows): # sheet.iter_rows()
                if i == 0 or i == 1 or i == 2: # skip first and second lines
                    continue
                f = False
                try:
                    g1 = row[3].value.encode('ascii','ignore').upper().strip()
                    g2 = row[6].value.encode('ascii','ignore').upper().strip()
                except:
                    f = True
                if f:
                    continue
                
                if g1.find('(') != -1:
                    t = g1.replace(')','').split('(')
                    g1 = [t[0].strip(),t[1].strip()]
                else:
                    g1 = [g1]

                if g2.find('(') != -1:
                    t = g2.replace(')','').split('(')
                    g2 = [t[0].strip(),t[1].strip()]
                else:
                    g2 = [g2]
                for gg1 in g1:
                    for gg2 in g2:
                        if gg1 and gg2 and gg1 != gg2:
                            (gg1,gg2) = (gg2,gg1) if gg2 < gg1 else (gg1,gg2)
                            data.add((gg1,gg2))

            print " - found",len(data),"fusions"

            # save version of
            txt = ['Prostate Tumor Patients (Robison et al. Cell 2015) database version: %s\n' % (today.strftime("%Y-%m-%d"),)]
            file(os.path.join(options.output_directory,'version.txt'),'a').writelines(txt)

    #
            # read the gene symbols
            file_symbols = os.path.join(options.output_directory,'synonyms.txt')
            loci = symbols.generate_loci(file_symbols)

            genes = symbols.read_genes_symbols(file_symbols)

            d = []
            for (g1,g2) in data:
                if g1.upper() != g2.upper():
                    ens1 = symbols.ensembl(g1.upper(),genes,loci)
                    ens2 = symbols.ensembl(g2.upper(),genes,loci)
                    if ens1 and ens2:
                        for e1 in ens1:
                            for e2 in ens2:
                                if e1 != e2:
                                    d.append([e1,e2])

            data = ['\t'.join(sorted(line)) + '\n' for line in d]
            data = sorted(set(data))

            print "%d known fusion genes found" % (len(data),)

            if not options.skip_filter_overlap:
                ens2hugo = dict([tuple(line.rstrip('\r\n').split('\t')) for line in file(os.path.join(options.output_directory,'genes_symbols.txt'),'r').readlines() if line.rstrip('\r\n')])

                d1 = []
                overlappings = ['ensembl_fully_overlapping_genes.txt',
                                'ensembl_same_strand_overlapping_genes.txt',
                                'gencode_fully_overlapping_genes.txt',
                                'gencode_same_strand_overlapping_genes.txt',
                                'refseq_fully_overlapping_genes.txt',
                                'refseq_same_strand_overlapping_genes.txt',
                                "ucsc_fully_overlapping_genes.txt",
                                "ucsc_same_strand_overlapping_genes.txt",
                                'pairs_pseudogenes.txt',
                                'banned.txt',
                                'dgd.txt',
                                'healthy.txt',
                                'paralogs.txt']
                ensembls = set(['ensembl_fully_overlapping_genes.txt',
                                'ensembl_same_strand_overlapping_genes.txt'])
                ens = []
                for ov in overlappings:
                    p = os.path.join(options.output_directory,ov)
                    print "Parsing file:",p
                    d2 = []
                    if os.path.isfile(p):
                        d2 = sorted(set([tuple(sorted(line.rstrip('\r\n').split('\t'))) for line in file(p,'r').readlines() if line.rstrip('\r\n')]))
                        d3 = set(['\t'.join(line)+'\n' for line in d2])
                        d4 = sorted([line for line in data if line in d3])
                        d4 = [line.rstrip('\r\n').split('\t') for line in d4]
                        d4 = [line+[ens2hugo.get(line[0],'')]+[ens2hugo.get(line[1],'')] for line in d4]
                        d4 = ['\t'.join(line)+'\n' for line in d4]
                        file(os.path.join(options.output_directory,"prostate_cancer___%s"%(ov,)),'w').writelines(d4)
                    if ov in ensembls:
                        ens.extend(d2)
                    d1.extend(d2)
                d = set()
                for line in d1:
                    (a,b) = (line[0],line[1])
                    if a > b:
                        (a,b) = (b,a)
                    d.add("%s\t%s\n" % (a,b))
                ens = set(['\t'.join(line)+'\n' for line in ens])
                ensembl = [line for line in data if line in ens]
                file(os.path.join(options.output_directory,'prostate_cancer__ensembl.txt'),'w').writelines(sorted(ensembl))
                skipped = [line for line in data if line in d]
                data = [line for line in data if line not in d]
                file(os.path.join(options.output_directory,'prostate_cancer__all.txt'),'w').writelines(sorted(skipped))

                print "%d known fusion genes left after removing the overlappings" % (len(data),)

        file(os.path.join(options.output_directory,'prostate_cancer.txt'),'w').writelines(data)

        if os.path.exists(tmp_file) and (not options.data_filename):
            os.remove(tmp_file)

    else:
        # write an empty file for other organisms than human
        file(os.path.join(options.output_directory,'prostate_cancer.txt'),'w').write('')
#
