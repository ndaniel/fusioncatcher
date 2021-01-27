#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""

It downloads the rt-circRNAs from article:
J.N. Vo, The Landscape of Circular RNA in Cancer, Cell, Feb. 2019, https://www.cell.com/cell/pdf/S0092-8674(18)31635-0.pdf



Author: Daniel Nicorici, Daniel.Nicorici@gmail.com

Copyright (c) 2009-2021 Daniel Nicorici

This file is part of FusionCatcher.

FusionCatcher is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

FusionCatcher is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with FusionCatcher (see file 'COPYING.txt').  If not, see
<http://www.gnu.org/licenses/>.

By default, FusionCatcher is running BLAT aligner
<http://users.soe.ucsc.edu/~kent/src/> but it offers also the option to disable
all its scripts which make use of BLAT aligner if you choose explicitly to do so.
BLAT's license does not allow to be used for commercial activities. If BLAT
license does not allow to be used in your case then you may still use
FusionCatcher by forcing not use the BLAT aligner by specifying the option
'--skip-blat'. Fore more information regarding BLAT please see its license.

Please, note that FusionCatcher does not require BLAT in order to find
candidate fusion genes!

This file is not running/executing/using BLAT.
"""



import os
import sys
import urllib2
import gzip
import socket
import optparse
import concatenate
import symbols
import shutil
import datetime
import zipfile

if __name__ == '__main__':

    #command line parsing

    usage = "%prog [options]"
    description = """It downloads the rt-circRNAs from article: J.N. Vo, The Landscape of Circular RNA in Cancer, Cell, Feb. 2019, https://www.cell.com/cell/pdf/S0092-8674(18)31635-0.pdf"""
    version = "%prog 0.12 beta"

    parser = optparse.OptionParser(usage=usage,description=description,version=version)

    parser.add_option("--organism","-g",
                      action = "store",
                      type = "string",
                      dest = "organism",
                      default = "homo_sapiens",
                      help="""The name of the organism for which the known fusion genes are downloaded, e.g. homo_sapiens, mus_musculus, etc. Default is '%default'.""")

    parser.add_option("--output","-o",
                      action="store",
                      type="string",
                      dest="output_directory",
                      default = '.',
                      help="""The output directory where the known fusion genes are stored. Default is '%default'.""")

    parser.add_option("--data","-d",
                      action="store",
                      type="string",
                      dest="data_filename",
                      help="""The input Excel file containg the data from the article. It should be used when there is no internet connection to the site which hosts the article.""")



    (options,args) = parser.parse_args()

    # validate options
    if not (options.output_directory
            ):
        parser.print_help()
        sys.exit(1)


    file(os.path.join(options.output_directory,'rtcircrnas.txt'),'w').write('')

    try:
        import openpyxl
        print "The Python Excel parser OPENPYXL was found!"
    except:
        print >> sys.stderr,"WARNING: Python OPENPYXL library was not found!"
        print >> sys.stderr,"Please, install the Python OpenPYXL library in order to be able to parse the known fusion genes in pancreatic tumor patients!"
        sys.exit(0)


    # timeout in seconds


    url_paper = "https://drive.google.com/file/d/166v5ZSm4x2ZtBmvoY3cNQRhtqq0nkbsO/view?usp=sharing"

    

    
    tmp_file = os.path.join(options.output_directory,'temp_rtcircrnas.xlsx')



    today = datetime.date.today()
    if options.organism.lower() == 'homo_sapiens':
        data = []

        if options.data_filename:
            tmp_file = options.data_filename
            print "Using the local file..."
        else:
            print "Downloading the known fusion genes from the article..."
            
            # get the URL first
            os.system("wget -S -O %s %s 2>&1" % (tmp_file,url_paper))

            
        data1 = []

        print "Parsing..."
        # parse the Excel file with the known fusion genes
        wb = None
        data1 = set()


        try:
            # new version of openpyxl
            print "Using the new version of openpyxl..."
            wb = openpyxl.load_workbook(filename=tmp_file)
        except:
            print "The new version of openpyxl did not work and now using the old version of openpyxl..."
            # old version of openpyxl
            try:
                wb = openpyxl.reader.excel.load_workbook(filename=tmp_file)
            except:
                print "WARNING: Not able to use the python openpyxl library! Aborting safely..."
                if not options.data_filename:
                    os.remove(tmp_file)
                sys.exit(0)

        print "..."
        # parse the Sheet1 sheet
        #print wb.get_sheet_names()
        sheet = wb.get_sheet_by_name(name = 'all rt_circRNA') #

        for i,row in enumerate(sheet.rows): # sheet.iter_rows()
            if i == 0: # skip first lines
                continue

            try:
                g1 = row[2].value.encode('ascii','ignore').upper().strip().strip("()")
                g2 = row[3].value.encode('ascii','ignore').upper().strip().strip("()")
                print g1,g2
            except:
                continue
            

            
            if g1 and g2 and g1 != g2:
                (g1,g2) = (g2,g1) if g2 < g1 else (g1,g2)
                data1.add((g1,g2))

        print " - found",len(data1)," rt-circ RNAs"



        # read the gene symbols
        file_symbols = os.path.join(options.output_directory,'synonyms.txt')
        loci = symbols.generate_loci(file_symbols)

        genes = symbols.read_genes_symbols(file_symbols)

        d = []
        for (g1,g2) in data1:
            if g1.upper() != g2.upper():
                ens1 = symbols.ensembl(g1.upper(),genes,loci)
                ens2 = symbols.ensembl(g2.upper(),genes,loci)
                if ens1 and ens2:
                    for e1 in ens1:
                        for e2 in ens2:
                            if e1 != e2:
                                d.append([e1,e2])

        data1 = ['\t'.join(sorted(line)) + '\n' for line in d]
        data1 = sorted(set(data1))

        print "%d known rt-circ RNAs" % (len(data1),)



        file(os.path.join(options.output_directory,'rtcircrnas.txt'),'w').writelines(data1)

        if os.path.exists(tmp_file) and (not options.data_filename):
            os.remove(tmp_file)


    else:
        # write an empty file for other organisms than human
        file(os.path.join(options.output_directory,'rtcircrnas.txt'),'w').write('')

        
    txt = ['rt-circ RNAs dataset (J.N. Vo, The Landscape of Circular RNA in Cancer, Cell, Feb. 2019): \n']
    file(os.path.join(options.output_directory,'version.txt'),'a').writelines(txt)
#
