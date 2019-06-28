#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""

It downloads the fusion genes from article:
Hu et al., TumorFusions: an integrative resource for cancer-associated transcript fusions, Nucleic Acids Research, Nov. 2017, https://doi.org/10.1093/nar/gkx1018



Author: Daniel Nicorici, Daniel.Nicorici@gmail.com

Copyright (c) 2009-2018 Daniel Nicorici

This file is part of FusionCatcher.

FusionCatcher is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

FusionCatcher is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with FusionCatcher (see file 'COPYING.txt').  If not, see
<http://www.gnu.org/licenses/>.

By default, FusionCatcher is running BLAT aligner
<http://users.soe.ucsc.edu/~kent/src/> but it offers also the option to disable
all its scripts which make use of BLAT aligner if you choose explicitly to do so.
BLAT's license does not allow to be used for commercial activities. If BLAT
license does not allow to be used in your case then you may still use
FusionCatcher by forcing not use the BLAT aligner by specifying the option
'--skip-blat'. Fore more information regarding BLAT please see its license.

Please, note that FusionCatcher does not require BLAT in order to find
candidate fusion genes!

This file is not running/executing/using BLAT.
"""



import os
import sys
import urllib2
import gzip
import socket
import optparse
import concatenate
import symbols
import shutil
import datetime
import zipfile

if __name__ == '__main__':

    #command line parsing

    usage = "%prog [options]"
    description = """It downloads the fusion genes from article: Hu et al., TumorFusions: an integrative resource for cancer-associated transcript fusions, Nucleic Acids Research, Nov. 2017, https://doi.org/10.1093/nar/gkx1018"""
    version = "%prog 0.12 beta"

    parser = optparse.OptionParser(usage=usage,description=description,version=version)

    parser.add_option("--organism","-g",
                      action = "store",
                      type = "string",
                      dest = "organism",
                      default = "homo_sapiens",
                      help="""The name of the organism for which the known fusion genes are downloaded, e.g. homo_sapiens, mus_musculus, etc. Default is '%default'.""")

    parser.add_option("--output","-o",
                      action="store",
                      type="string",
                      dest="output_directory",
                      default = '.',
                      help="""The output directory where the known fusion genes are stored. Default is '%default'.""")

    parser.add_option("--data",
                      action="store",
                      type="string",
                      dest="data_filename",
                      help="""The input Excel file containg the data from the article. It should be used when there is no internet connection to the site which hosts the article.""")


    parser.add_option("--skip-filter-overlap",
                      action="store_true",
                      dest="skip_filter_overlap",
                      default = False,
                      help="""If set then it filters out the known fusion genes where the (i) genes are fully overlapping, or (ii) the genes are partially overlapping and are on the same strand. Default is '%default'.""")


    (options,args) = parser.parse_args()

    # validate options
    if not (options.output_directory
            ):
        parser.print_help()
        sys.exit(1)


    file(os.path.join(options.output_directory,'tcga-cancer.txt'),'w').write('')
    file(os.path.join(options.output_directory,'tcga-normal.txt'),'w').write('')

    try:
        import openpyxl
        print "The Python Excel parser OPENPYXL was found!"
    except:
        print >> sys.stderr,"WARNING: Python OPENPYXL library was not found!"
        print >> sys.stderr,"Please, install the Python OpenPYXL library in order to be able to parse the known fusion genes in pancreatic tumor patients!"
        sys.exit(0)


    # timeout in seconds
    timeout = 1800
    socket.setdefaulttimeout(timeout)

    url_paper = "https://academic.oup.com/nar/article/doi/10.1093/nar/gkx1018/4584571"
    url_file = os.path.join(options.output_directory,'temp_url.txt')
    

    #url = 'https://oup.silverchair-cdn.com/oup/backfile/Content_public/Journal/nar/PAP/10.1093_nar_gkx1018/1/gkx1018_supp.zip'
    url = "https://oup.silverchair-cdn.com/oup/backfile/Content_public/Journal/nar/PAP/10.1093_nar_gkx1018/1/gkx1018_supp.zip"
    
    tmp_file = os.path.join(options.output_directory,'temp_tumorfusions.zip')
    tmp2_file = os.path.join(options.output_directory,'temp_tumorfusions.xlsx')

#    headers = { 'User-Agent' : 'Mozilla/5.0' }
    headers = {     'User-agent': 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-GB; rv:1.9.2.3) Gecko/20100401 Firefox/3.6.3',
    'Accept' : 'text/xml,application/xml,application/xhtml+xml,text/html;q=0.9,text/plain;q=0.8,image/png,*/*;q=0.5',
    'Accept-Language' : 'en-gb,en;q=0.5'
    }

    today = datetime.date.today()
    if options.organism.lower() == 'homo_sapiens':
        data = []
        sem = True
        if options.data_filename:
            tmp_file = options.data_filename
            print "Using the local file..."
        else:
            print "Downloading the known fusion genes from the article..."
            
            # get the URL first
            os.system("wget -S -O - %s 2>&1| grep 'Supplementary Data' | grep NAR > %s" % (url_paper,url_file))
            if os.path.exists(url_file):
                rx = file(url_file,"r").readline()
                if rx:
                    ux = rx.split('href="')[1].split('">')[0]
                    if ux:
                        url = ux
                        #print url
                
            
            f = True
            try:
                req = urllib2.Request('%s' % (url,), headers=headers)
                da = urllib2.urlopen(req)
                file(tmp_file,'w').write(da.read())
            except:
                f = False
                print >>sys.stderr, "Warning: Cannot access '%s'! The output file will be empty!" % (url,)
                sem = False
            if f:
                print "Downloaded succesfully!"

        if not sem:
            # write an empty file for other organisms than human
            file(os.path.join(options.output_directory,'tcga-cancer.txt'),'w').write('')
            file(os.path.join(options.output_directory,'tcga-normal.txt'),'w').write('')
            txt = ['TumorFusions (Hu et al. Nucleic Acids Research) database version: Nov. 2017\n']
            file(os.path.join(options.output_directory,'version.txt'),'a').writelines(txt)
            sys.exit(0)
            
        zf = zipfile.ZipFile(tmp_file)
        file(tmp2_file,'w').write(zf.read('nar-02671-data-e-2017-File007.xlsx'))
        zf.close()
        data1 = []
        data2 = []
        if sem:
            print "Parsing..."
            # parse the Excel file with the known fusion genes
            wb = None
            data1 = set()
            data2 = set()

            try:
                # new version of openpyxl
                print "Using the new version of openpyxl..."
                wb = openpyxl.load_workbook(filename=tmp2_file)
            except:
                print "The new version of openpyxl did not work and now using the old version of openpyxl..."
                # old version of openpyxl
                try:
                    wb = openpyxl.reader.excel.load_workbook(filename=tmp_file)
                except:
                    print "WARNING: Not able to use the python openpyxl library!"
                    if not options.data_filename:
                        os.remove(tmp_file)
                        os.remove(tmp2_file)
                    sys.exit(0)

            print "..."
            # parse the Sheet1 sheet
            #print wb.get_sheet_names()
            sheet = wb.get_sheet_by_name(name = 'Cancer fusions') #

            for i,row in enumerate(sheet.rows): # sheet.iter_rows()
                if i == 0: # skip first and second lines
                    continue
                f = False
                try:
                    g1 = row[2].value.encode('ascii','ignore').upper().strip()
                    g2 = row[3].value.encode('ascii','ignore').upper().strip()
                except:
                    f = True
                if f:
                    continue
                
                if g1 and g2 and g1 != g2:
                    (g1,g2) = (g2,g1) if g2 < g1 else (g1,g2)
                    data1.add((g1,g2))

            print " - found",len(data1)," cancer fusions"


            sheet = wb.get_sheet_by_name(name = 'Normal tissue fusions') #

            for i,row in enumerate(sheet.rows): # sheet.iter_rows()
                if i == 0: # skip first and second lines
                    continue
                f = False
                try:
                    g1 = row[2].value.encode('ascii','ignore').upper().strip()
                    g2 = row[3].value.encode('ascii','ignore').upper().strip()
                except:
                    f = True
                if f:
                    continue
                
                if g1 and g2 and g1 != g2:
                    (g1,g2) = (g2,g1) if g2 < g1 else (g1,g2)
                    data2.add((g1,g2))

            print " - found",len(data2),"non-cancer fusions"

            # save version of

    #
            # read the gene symbols
            file_symbols = os.path.join(options.output_directory,'synonyms.txt')
            loci = symbols.generate_loci(file_symbols)

            genes = symbols.read_genes_symbols(file_symbols)

            d = []
            for (g1,g2) in data1:
                if g1.upper() != g2.upper():
                    ens1 = symbols.ensembl(g1.upper(),genes,loci)
                    ens2 = symbols.ensembl(g2.upper(),genes,loci)
                    if ens1 and ens2:
                        for e1 in ens1:
                            for e2 in ens2:
                                if e1 != e2:
                                    d.append([e1,e2])

            data1 = ['\t'.join(sorted(line)) + '\n' for line in d]
            data1 = sorted(set(data1))

            print "%d known cancer fusion genes found" % (len(data1),)

            d = []
            for (g1,g2) in data2:
                if g1.upper() != g2.upper():
                    ens1 = symbols.ensembl(g1.upper(),genes,loci)
                    ens2 = symbols.ensembl(g2.upper(),genes,loci)
                    if ens1 and ens2:
                        for e1 in ens1:
                            for e2 in ens2:
                                if e1 != e2:
                                    d.append([e1,e2])

            data2 = ['\t'.join(sorted(line)) + '\n' for line in d]
            data2 = sorted(set(data2))

            print "%d known non-cancer fusion genes found" % (len(data2),)

            if not options.skip_filter_overlap:
                ens2hugo = dict([tuple(line.rstrip('\r\n').split('\t')) for line in file(os.path.join(options.output_directory,'genes_symbols.txt'),'r').readlines() if line.rstrip('\r\n')])

                d1 = []
                overlappings = ['ensembl_fully_overlapping_genes.txt',
                                'ensembl_same_strand_overlapping_genes.txt',
                                'gencode_fully_overlapping_genes.txt',
                                'gencode_same_strand_overlapping_genes.txt',
                                'refseq_fully_overlapping_genes.txt',
                                'refseq_same_strand_overlapping_genes.txt',
                                "ucsc_fully_overlapping_genes.txt",
                                "ucsc_same_strand_overlapping_genes.txt",
                                'pairs_pseudogenes.txt',
                                'banned.txt',
                                'dgd.txt',
                                'healthy.txt',
                                'paralogs.txt']
                ensembls = set(['ensembl_fully_overlapping_genes.txt',
                                'ensembl_same_strand_overlapping_genes.txt'])
                ens = []
                for ov in overlappings:
                    p = os.path.join(options.output_directory,ov)
                    print "Parsing file:",p
                    d2 = []
                    if os.path.isfile(p):
                        d2 = sorted(set([tuple(sorted(line.rstrip('\r\n').split('\t'))) for line in file(p,'r').readlines() if line.rstrip('\r\n')]))
                        d3 = set(['\t'.join(line)+'\n' for line in d2])
                        d4 = sorted([line for line in data1 if line in d3])
                        d4 = [line.rstrip('\r\n').split('\t') for line in d4]
                        d4 = [line+[ens2hugo.get(line[0],'')]+[ens2hugo.get(line[1],'')] for line in d4]
                        d4 = ['\t'.join(line)+'\n' for line in d4]
                        file(os.path.join(options.output_directory,"tcga-2___%s"%(ov,)),'w').writelines(d4)
                    if ov in ensembls:
                        ens.extend(d2)
                    d1.extend(d2)
                d = set()
                for line in d1:
                    (a,b) = (line[0],line[1])
                    if a > b:
                        (a,b) = (b,a)
                    d.add("%s\t%s\n" % (a,b))
                ens = set(['\t'.join(line)+'\n' for line in ens])
                ensembl = [line for line in data1 if line in ens]
                file(os.path.join(options.output_directory,'tcga-2___ensembl.txt'),'w').writelines(sorted(ensembl))
                skipped = [line for line in data1 if line in d]
                data1 = [line for line in data1 if line not in d]
                file(os.path.join(options.output_directory,'tcga-2___all.txt'),'w').writelines(sorted(skipped))

                print "%d known fusion genes left after removing the overlappings" % (len(data1),)

        file(os.path.join(options.output_directory,'tcga-cancer.txt'),'w').writelines(data1)
        file(os.path.join(options.output_directory,'tcga-normal.txt'),'w').writelines(data2)

        if os.path.exists(tmp_file) and (not options.data_filename):
            os.remove(tmp_file)
            os.remove(tmp2_file)

    else:
        # write an empty file for other organisms than human
        file(os.path.join(options.output_directory,'tcga-cancer.txt'),'w').write('')
        file(os.path.join(options.output_directory,'tcga-normal.txt'),'w').write('')
        
    txt = ['TumorFusions (Hu et al. Nucleic Acids Research) database version: Nov. 2017\n']
    file(os.path.join(options.output_directory,'version.txt'),'a').writelines(txt)
#
