#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
It downloads the lastest known conjoined genes from the ConjoinG database
<http://metasystems.riken.jp/conjoing/>. It takes only the conjoined genes
from the file:
"http://metasystems.riken.jp/conjoing/download/SuppTable_CG_Tissues.xlsx".



Author: Daniel Nicorici, Daniel.Nicorici@gmail.com

Copyright (c) 2009-2022 Daniel Nicorici

This file is part of FusionCatcher.

FusionCatcher is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

FusionCatcher is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with FusionCatcher (see file 'COPYING.txt').  If not, see
<http://www.gnu.org/licenses/>.

By default, FusionCatcher is running BLAT aligner
<http://users.soe.ucsc.edu/~kent/src/> but it offers also the option to disable
all its scripts which make use of BLAT aligner if you choose explicitly to do so.
BLAT's license does not allow to be used for commercial activities. If BLAT
license does not allow to be used in your case then you may still use
FusionCatcher by forcing not use the BLAT aligner by specifying the option
'--skip-blat'. Fore more information regarding BLAT please see its license.

Please, note that FusionCatcher does not require BLAT in order to find
candidate fusion genes!

This file is not running/executing/using BLAT.
"""



import os
import sys
import socket
import optparse
import shutil
import urllib2
import symbols
import datetime


if __name__ == '__main__':

    #command line parsing

    usage = "%prog [options]"
    description = """It downloads the lastest known conjoined genes from the ConjoinG database
<http://metasystems.riken.jp/conjoing/>."""
    version = "%prog 0.10 beta"

    parser = optparse.OptionParser(usage=usage,description=description,version=version)

    parser.add_option("--organism",
                      action = "store",
                      type = "string",
                      dest = "organism",
                      default = "homo_sapiens",
                      help="""The name of the organism for which the known conjoined genes are downloaded, e.g. homo_sapiens, mus_musculus, etc. Default is '%default'.""")

    parser.add_option("--output",
                      action="store",
                      type="string",
                      dest="output_directory",
                      default = '.',
                      help="""The output directory where the known fusion genes are stored. Default is '%default'.""")

    parser.add_option("--server",
                      action="store",
                      type="string",
                      dest="server",
                      default = "http://metasystems.riken.jp",
                      help="""The ConjoinG server from where the known conjoined genes are downloaded. Default is '%default'.""")

    (options,args) = parser.parse_args()

    # validate options
    if not (options.output_directory
            ):
        parser.print_help()
        sys.exit(1)

    try:
        import openpyxl
        print "The Python Excel parser OPENPYXL was found!"
    except:
        print >> sys.stderr,"WARNING: Python OPENPYXL library was not found!"
        print >> sys.stderr,"Please, install the Python OpenPYXL library in order to be able to parse the ConjoinG database!"
        file(os.path.join(options.output_directory,'conjoing.txt'),'w').write('')
        sys.exit(0)


    # timeout in seconds
    timeout = 1800
    socket.setdefaulttimeout(timeout)
    tmp_file = 'temp_conjoing.xlsx'
    tmp2_file = 'temp_conjoing2.xlsx'
    tmp3_file = 'temp_conjoing3.txt'

    #http://metasystems.riken.jp/conjoing/download/SuppTable_CG_Tissues.xlsx
    url = '/conjoing/download/SuppTable_CG_Tissues.xlsx'
    url2 = '/conjoing/download/SuppTable_ParentGeneDistance.xlsx'
    url3 = '/conjoing/result'

    headers = { 'User-Agent' : 'Mozilla/5.0' }

    if options.organism.lower() == 'homo_sapiens':
        today = datetime.date.today()
        file(os.path.join(options.output_directory,'version.txt'),'a').writelines(['ConjoinG database version: %s\n' % (today.strftime("%Y-%m-%d"),)])

        print "Downloading the known conjoined genes from ConjoinG database!"

        sem = True
        try:
            req = urllib2.Request('%s%s' % (options.server,url), headers=headers)
            d = urllib2.urlopen(req)
            file(tmp_file,'w').write(d.read())
        except:
            sem = False
            print >>sys.stderr, "Warning: Cannot access '%s%s'! The output file will be empty!" % (options.server,url)

        try:
            req = urllib2.Request('%s%s' % (options.server,url2), headers=headers)
            d = urllib2.urlopen(req)
            file(tmp2_file,'w').write(d.read())
        except:
            sem = False
            print >>sys.stderr, "Warning: Cannot access '%s%s'! The output file will be empty!" % (options.server,url2)

        try:
            req = urllib2.Request('%s%s' % (options.server,url3), headers=headers)
            d = urllib2.urlopen(req)
            file(tmp3_file,'w').write(d.read())
        except:
            sem = False
            print >>sys.stderr, "Warning: Cannot access '%s%s'! The output file will be empty!" % (options.server,url3)

        if sem:
            print "Parsing..."
            # parse the Conjoing file with the known conjoing genes
            try:
                # new version of openpyxl
                print "Using the new version of openpyxl..."
                wb = openpyxl.load_workbook(filename=tmp_file)
                wb2 = openpyxl.load_workbook(filename=tmp2_file)
            except:
                print "The new version of openpyxl did not work and now using the old version of openpyxl..."
                # old version of openpyxl
                try:
                    wb = openpyxl.reader.excel.load_workbook(filename=tmp_file)
                    wb2 = openpyxl.reader.excel.load_workbook(filename=tmp2_file)
                except:
                    print "WARNING: Not able to use the python openpyxl library!"
                    file(os.path.join(options.output_directory,'conjoing.txt'),'w').write('')
                    os.remove(tmp_file)
                    os.remove(tmp2_file)
                    os.remove(tmp3_file)
                    sys.exit(0)

            fusions = set()

            # parse the Sheet1 sheet
            sheet = wb.get_sheet_by_name(name = 'Sheet1') #
            sheet2 = wb2.get_sheet_by_name(name = 'Sheet1') #

            try:
                for i,row in enumerate(sheet.rows): # sheet.iter_rows()
                    if i == 0 or i == 1: # skip first and second lines
                        continue
                    gg = row[1].value
                    gg = gg.upper().encode('ascii','ignore').split('-')
                    g1 = gg[0]
                    g2 = gg[1]
                    (g1,g2) = (g2,g1) if g2 < g1 else (g1,g2)
                    fusions.add((g1,g2))

                for i,row in enumerate(sheet2.rows): # sheet.iter_rows()
                    if i == 0 or i == 1: # skip first and second lines
                        continue
                    g1 = str(row[2].value).encode('ascii','ignore')
                    g2 = str(row[3].value).encode('ascii','ignore')
                    g1 = g1.upper()
                    g2 = g2.upper()
                    (g1,g2) = (g2,g1) if g2 < g1 else (g1,g2)
                    fusions.add((g1,g2))

                # parse the HTML file
                html = [line.rstrip('\r\n') for line in file(tmp3_file,'r').readlines() if line.lower().find("prime5 gene")!=-1 ]
                html = [line.split(' gene"')[1:] for line in html]
                for line in html:
                    row = [el.split('<')[0].upper() for el in line]

                    row = [el[1:] if el.startswith('>') else el for el in row]
                    if row:
                        row = list(set(row))
                        n = len(row)

                        for i in xrange(n-1):
                            for j in xrange(i+1,n):
                                g1 = row[i]
                                g2 = row[j]
                                (g1,g2) = (g2,g1) if g2 < g1 else (g1,g2)
                                fusions.add((g1,g2))

                fusions = list(fusions)
                print "%d known conjoined genes found!" % (len(fusions),)

                # read the gene symbols
                file_symbols = os.path.join(options.output_directory,'synonyms.txt')
                genes = symbols.read_genes_symbols(file_symbols)

                banned = set()
                loci = symbols.generate_loci(file_symbols)
                #for v in symbols.locus.values():
                for v in loci.values():
                    if v:
                        n = len(v)
                        if n > 1:
                            for i in xrange(n-1):
                                for j in xrange(i+1,n):
                                    if v[i].upper() != v[j].upper():
                                        ens1 = symbols.ensembl(v[i].upper(),genes,loci)
                                        ens2 = symbols.ensembl(v[j].upper(),genes,loci)
                                        if ens1 and ens2:
                                            for e1 in ens1:
                                                for e2 in ens2:
                                                    if e1 != e2:
                                                        (e1,e2) = (e2,e1) if e2 < e1 else (e1,e2)
                                                        banned.add((e1,e2))


                d = []
                for (g1,g2) in fusions:
                    if ( g1.upper() == g2.upper() or ((g1.endswith('@') and g2.endswith('@')) and g1.upper()[:2] == g2.upper()[:2])):
                        print "%s-%s skipped!" % (g1,g2)
                        continue
                    ens1 = symbols.ensembl(g1,genes,loci)
                    ens2 = symbols.ensembl(g2,genes,loci)

                    if ens1 and ens2:
                        for e1 in ens1:
                            for e2 in ens2:
                                if e1 != e2 and ((e1,e2) not in banned) and ((e2,e1) not in banned):
                                    d.append([e1,e2])


                data = ['\t'.join(sorted(line)) + '\n' for line in d]
                data = sorted(set(data))

                print "%d known conjoined genes converted succesfully to Ensembl Gene ids!" % (len(data),)
            except:
                data = []
                print "Ups..."
                print "WARNING: Not able to parse the ConjoinG database!"
        else:
            data = []

        file(os.path.join(options.output_directory,'conjoing.txt'),'w').writelines(data)
        if os.path.isfile(tmp_file):
            os.remove(tmp_file)
        if os.path.isfile(tmp2_file):
            os.remove(tmp2_file)
        if os.path.isfile(tmp3_file):
            os.remove(tmp3_file)
    else:
        # write an empty file for other organisms than human
        file(os.path.join(options.output_directory,'conjoing.txt'),'w').write('')
#
